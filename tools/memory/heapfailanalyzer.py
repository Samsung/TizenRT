#!/usr/bin/env python3
###########################################################################
#
# Copyright 2025 Samsung Electronics All Rights Reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing,
# software distributed under the License is distributed on an
# "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND,
# either express or implied. See the License for the specific
# language governing permissions and limitations under the License.
#
###########################################################################
# File : heapfailanalyzer.py
# Description: Tool to parse the memory failure log and provides a summary of memory usage by PID and owner.

"""
TizenRT Memory Allocation Failure Log Analyzer
----------------------------------------------
This script parses TizenRT memory allocation failure logs and provides a summary
of memory usage by PID and owner using a two-pass approach for better performance.
Enhanced to handle logs with missing newlines, timestamps, and multiple entries per line.
"""

import re
import sys
import argparse
from collections import defaultdict
import datetime
import os
import time
import csv
from openpyxl import Workbook

def preprocess_log_file(input_filename, output_filename=None):
    """
    Preprocesses the log file to fix missing newlines and normalize format
    - Handles multiple log entries on a single line
    - Preserves timestamps when available
    - Adds missing timestamps when needed
    - Creates a cleaner normalized format for easier parsing
    
    Args:
        input_filename (str): Path to the input log file
        output_filename (str, optional): Path to output the preprocessed log file
                                         If None, a temporary file will be created
    
    Returns:
        str: Path to the preprocessed file
    """
    if output_filename is None:
        output_filename = input_filename + ".processed"
    
    try:
        print(f"Preprocessing log file to handle inconsistent formatting...")
        
        with open(input_filename, 'r', encoding='utf-8', errors='replace') as infile, \
             open(output_filename, 'w', encoding='utf-8') as outfile:
            
            # Read the whole file content
            content = infile.read()
            
            # Split multiple entries that appear on the same line by inserting newlines
            # Look for timestamp pattern [YYYY-MM-DD HH:MM:SS.mmm] inside a line
            content = re.sub(r'(\]\s*[^\[\n]+)(\[\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}\.\d+\])', r'\1\n\2', content)
            
            # Also split entries that have no timestamp but follow a pattern
            content = re.sub(r'(\|[^|]*\|)(\s*0x[0-9a-f]+\s*\|)', r'\1\n[MISSING_TIMESTAMP] \2', content)
            
            # Fix potentially merged memory entries that should be on separate lines
            # Also handle the case where timestamp is missing entirely
            content = re.sub(r'((?:mm_manage_alloc_fail|mm_manage_alloc_fail_dump):.*?[^0-9a-f])(\s*(?:mm_manage_alloc_fail|mm_manage_alloc_fail_dump):)', r'\1\n\2', content)
            
            # Add missing timestamps if needed for entries without any timestamp
            content = re.sub(r'(^|\n)(?!\[)(\s*)(0x[0-9a-f]+\s*\|)', r'\1[MISSING_TIMESTAMP] \3', content)
            content = re.sub(r'(^|\n)(?!\[)(\s*)((?:mm_manage_alloc_fail|mm_manage_alloc_fail_dump):)', r'\1[MISSING_TIMESTAMP] \3', content)
            
            # Write the processed content
            outfile.write(content)
        
        print(f"Preprocessing complete.")
        return output_filename
        
    except Exception as e:
        print(f"Error preprocessing log file: {str(e)}")
        return input_filename  # Return original filename if preprocessing fails

def count_failures(filename):
    """
    First pass: Count failures and measure file stats for better progress estimation
   
    Args:
        filename (str): Path to the log file
       
    Returns:
        tuple: (num_failures, file_size, line_markers)
    """
    # Pattern that can handle missing or malformed timestamps and multiple entries per line
    failure_start_pattern = re.compile(r'(?:\[([\d\-: .]+)\])?\s*(?:mm_manage_alloc_fail|mm_manage_alloc_fail_dump): Allocation failed from user heap\.')
   
    failure_count = 0
    file_size = os.path.getsize(filename)
   
    # Store positions of each 10th failure for better progress tracking
    failure_positions = []
   
    try:
        print(f"First pass: Analyzing log file structure ({format_size(file_size)})...")
        print("[" + " " * 50 + "] 0%", end="\r")
       
        with open(filename, 'r', encoding='utf-8', errors='replace') as f:
            bytes_read = 0
            last_progress = 0
            buffer_size = 4 * 1024 * 1024  # 4MB buffer
           
            # Using line-based processing for the first pass to count failures
            buffer = ""
            for chunk in iter(lambda: f.read(buffer_size), ''):
                chunk_size = len(chunk.encode('utf-8'))  # Get encoded size
                bytes_read += chunk_size
               
                # Add current chunk to buffer
                buffer += chunk
                
                # Count failure starts in this buffer
                # This will find ALL matches in the buffer, even if they're on the same line
                matches = list(failure_start_pattern.finditer(buffer))
                failure_count += len(matches)
                
                # Store positions of failures for progress tracking
                for match in matches:
                    if failure_count % 10 == 0:
                        position = bytes_read - len(buffer) + match.start()
                        failure_positions.append((failure_count, position))
                
                # Keep only the last part of the buffer to avoid missing matches that span chunks
                last_newline = buffer.rfind('\n')
                if last_newline > 0:
                    buffer = buffer[last_newline:]
               
                # Update progress bar
                progress = int(50 * bytes_read / file_size)
                if progress > last_progress:
                    print(f"[{'#' * progress}{' ' * (50 - progress)}] {int(100 * bytes_read / file_size)}%", end="\r")
                    last_progress = progress
       
        print(f"\nFound {failure_count} memory allocation failures.")
       
    except FileNotFoundError:
        print(f"Error: File '{filename}' not found")
        sys.exit(1)
    except Exception as e:
        print(f"Error analyzing log file: {str(e)}")
        sys.exit(1)
   
    return failure_count, file_size, failure_positions

def parse_log_file(filename, total_failures, file_size, failure_positions):
    """
    Second pass: Parse the log file and extract memory allocation failure information
    with better progress tracking. Now handles multiple entries per line.
   
    Args:
        filename (str): Path to the log file
        total_failures (int): Total failures counted in first pass
        file_size (int): Total file size
        failure_positions (list): Positions of failures for tracking
       
    Returns:
        dict: Summary dictionary
    """
    # Initialize summary 
    summary = {
        'total_failures': total_failures,
        'by_pid': defaultdict(lambda: {'count': 0, 'total_allocated': 0}),
        'by_owner': defaultdict(lambda: {'count': 0, 'total_allocated': 0, 'pids': set()}),
        'largest_allocation': {'size': 0, 'pid': None, 'owner': None, 'address': None},
        'total_requested_size': 0,
        'requested_sizes': [],  # For calculating average later
        'total_allocated_memory': 0,
        'total_free_memory': 0,
        'top_pids': [],  # Updated to store top PIDs
        'top_owners': []  # Updated to store top owners
    }
   
    # Precompile regex patterns - enhanced for better matching across lines
    failure_start_pattern = re.compile(r'(?:\[([\d\-: .]+)\])?\s*(?:mm_manage_alloc_fail|mm_manage_alloc_fail_dump): Allocation failed from user heap\.')
    size_pattern = re.compile(r'(?:\[.*?\])?\s*(?:mm_manage_alloc_fail|mm_manage_alloc_fail_dump):  - requested size (\d+)')
    caller_pattern = re.compile(r'(?:\[.*?\])?\s*(?:mm_manage_alloc_fail|mm_manage_alloc_fail_dump):  - caller address = (0x[0-9a-f]+)')
    largest_free_pattern = re.compile(r'(?:\[.*?\])?\s*(?:mm_manage_alloc_fail|mm_manage_alloc_fail_dump):  - largest free size : (\d+)')
    total_free_pattern = re.compile(r'(?:\[.*?\])?\s*(?:mm_manage_alloc_fail|mm_manage_alloc_fail_dump):  - total free size   : (\d+)')
    header_pattern = re.compile(r'(?:\[.*?\])?\s*MemAddr \|   Size   \| Status \|    Owner   \|  Pid  \|')
    entry_pattern = re.compile(r'(?:\[.*?\])?\s*(0x[0-9a-f]+)\s*\|\s*(\d+)\s*\|\s*([A-Z])\s*\|\s*(0x[0-9a-f ]+)\s*\|\s*(\d+)(?:\(([A-Z])\))?\s*\|')
   
    # Initialize processing variables
    current_failure = None
    heap_entries = []
    failures_processed = 0
    next_marker_idx = 0
    in_table = False
   
    # Create failure marker map for better progress tracking
    failure_markers = {}
    if failure_positions:
        for count, position in failure_positions:
            failure_markers[count] = position
   
    try:
        print(f"Second pass: Processing memory allocation data...")
        print("[" + " " * 50 + "] 0%", end="\r")
       
        start_time = time.time()
        last_update_time = start_time
        bytes_read = 0
        last_progress = 0
        
        # Read the file content
        with open(filename, 'r', encoding='utf-8', errors='replace') as f:
            content = f.read()
            
        # First find all failure start positions in the content
        failure_starts = []
        for match in failure_start_pattern.finditer(content):
            timestamp = match.group(1) if match.group(1) else "MISSING_TIMESTAMP"
            position = match.start()
            failure_starts.append((position, timestamp))
        
        # Sort by position to ensure we process them in order
        failure_starts.sort()
        
        # Add an end marker
        failure_starts.append((len(content), None))
        
        # Process each failure section
        for i in range(len(failure_starts) - 1):
            start_pos, timestamp = failure_starts[i]
            end_pos = failure_starts[i+1][0]
            
            # Get failure section content
            section = content[start_pos:end_pos]
            
            # Process this failure
            current_failure = {
                'timestamp': timestamp,
                'requested_size': None,
                'caller_address': None,
                'largest_free_size': None,
                'total_free_size': None
            }
            
            # Extract information from this section
            size_match = size_pattern.search(section)
            if size_match:
                current_failure['requested_size'] = int(size_match.group(1))
            
            caller_match = caller_pattern.search(section)
            if caller_match:
                current_failure['caller_address'] = caller_match.group(1)
            
            largest_free_match = largest_free_pattern.search(section)
            if largest_free_match:
                current_failure['largest_free_size'] = int(largest_free_match.group(1))
            
            total_free_match = total_free_pattern.search(section)
            if total_free_match:
                current_failure['total_free_size'] = int(total_free_match.group(1))
            
            # Find memory table entries in this section
            if header_pattern.search(section):
                # Parse all entries in this section
                for entry_match in entry_pattern.finditer(section):
                    address = entry_match.group(1)
                    size = int(entry_match.group(2))
                    status = entry_match.group(3)
                    owner = entry_match.group(4).strip()
                    pid = entry_match.group(5)
                    special = entry_match.group(6) if entry_match.group(6) else ""
                    
                    # Update summary directly
                    summary['by_pid'][pid]['count'] += 1
                    summary['by_pid'][pid]['total_allocated'] += size
                    
                    summary['by_owner'][owner]['count'] += 1
                    summary['by_owner'][owner]['total_allocated'] += size
                    summary['by_owner'][owner]['pids'].add(pid)
                    
                    summary['total_allocated_memory'] += size
                    
                    # Check for largest allocation
                    if size > summary['largest_allocation']['size']:
                        summary['largest_allocation'] = {
                            'size': size,
                            'pid': pid,
                            'owner': owner,
                            'address': address
                        }
            
            # Update requested sizes for this failure
            if current_failure.get('requested_size'):
                summary['total_requested_size'] += current_failure['requested_size']
                summary['requested_sizes'].append(current_failure['requested_size'])
            
            # Calculate progress and update display
            failures_processed += 1
            progress = int(50 * failures_processed / total_failures) if total_failures > 0 else 0
            current_time = time.time()
            
            if progress > last_progress or current_time - last_update_time > 1.0:
                elapsed_time = current_time - start_time
                
                # Calculate ETA
                if failures_processed > 0 and total_failures > 0:
                    estimated_total = (elapsed_time / (failures_processed / total_failures))
                    remaining_time = max(0, estimated_total - elapsed_time)
                    eta_text = f"ETA: {format_time(remaining_time)}"
                else:
                    eta_text = "Calculating ETA..."
                
                print(f"[{'#' * progress}{' ' * (50 - progress)}] {failures_processed}/{total_failures} failures ({int(100 * failures_processed / total_failures)}%) - {eta_text}", end="\r")
                last_progress = progress
                last_update_time = current_time
       
        # Calculate average requested size
        if summary['requested_sizes']:
            summary['average_requested_size'] = sum(summary['requested_sizes']) / len(summary['requested_sizes'])
        else:
            summary['average_requested_size'] = 0
       
        # Find top 5 PIDs and owners by memory allocation
        sorted_pids = sorted(summary['by_pid'].items(), key=lambda x: x[1]['total_allocated'], reverse=True)
        sorted_owners = sorted(summary['by_owner'].items(), key=lambda x: x[1]['total_allocated'], reverse=True)
       
        # Get top 5 PIDs or all if less than 5
        num_pids = min(5, len(sorted_pids))
        for i in range(num_pids):
            pid, data = sorted_pids[i]
            summary['top_pids'].append({
                'pid': pid,
                'count': data['count'],
                'total_allocated': data['total_allocated']
            })
       
        # Get top 5 owners or all if less than 5
        num_owners = min(5, len(sorted_owners))
        for i in range(num_owners):
            owner, data = sorted_owners[i]
            summary['top_owners'].append({
                'owner': owner,
                'count': data['count'],
                'total_allocated': data['total_allocated'],
                'pids': data['pids']
            })
       
        total_time = time.time() - start_time
        print(f"\nProcessing completed in {format_time(total_time)}")
       
    except Exception as e:
        print(f"\nError processing log file: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
   
    return summary

def format_size(size_bytes):
    """Format file size in human-readable format"""
    if size_bytes < 0:
        return "0 B"
   
    for unit in ['B', 'KB', 'MB', 'GB']:
        if size_bytes < 1024.0 or unit == 'GB':
            break
        size_bytes /= 1024.0
    return f"{size_bytes:.2f} {unit}"

def format_time(seconds):
    """Format time in human-readable format"""
    if seconds < 60:
        return f"{seconds:.1f}s"
    elif seconds < 3600:
        minutes = int(seconds // 60)
        seconds = int(seconds % 60)
        return f"{minutes}m {seconds}s"
    else:
        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        return f"{hours}h {minutes}m"

def print_summary(summary):
    """
    Print a formatted summary of memory allocation failures
   
    Args:
        summary (dict): Summary statistics
    """
    print("\n===== MEMORY ALLOCATION FAILURE SUMMARY =====\n")
    print(f"Total failures analyzed: {summary['total_failures']}")
    print(f"Total requested memory: {summary['total_requested_size']} bytes ({format_size(summary['total_requested_size'])})")
    print(f"Average request size: {summary['average_requested_size']:.2f} bytes ({format_size(summary['average_requested_size'])})")
    print(f"Total allocated memory: {summary['total_allocated_memory']} bytes ({format_size(summary['total_allocated_memory'])})")
   
    if summary['largest_allocation']['size'] > 0:
        print(f"\nLargest allocation:")
        print(f"  Size: {summary['largest_allocation']['size']} bytes ({format_size(summary['largest_allocation']['size'])})")
        print(f"  PID: {summary['largest_allocation']['pid']}")
        print(f"  Owner: {summary['largest_allocation']['owner']}")
        print(f"  Address: {summary['largest_allocation']['address']}")
   
    print("\n----- PID SUMMARY -----")
    print(f"{'PID':<8} {'Count':<8} {'Total Size':<20} {'% of Memory':<12}")
    print("-" * 60)
   
    # Sort PIDs by total allocated memory
    sorted_pids = sorted(summary['by_pid'].items(), key=lambda x: x[1]['total_allocated'], reverse=True)
   
    for pid, data in sorted_pids:
        percent = (data['total_allocated'] / summary['total_allocated_memory'] * 100) if summary['total_allocated_memory'] else 0
        print(f"{pid:<8} {data['count']:<8} {data['total_allocated']:<12} ({format_size(data['total_allocated']):<6}) {percent:<12.2f}")
   
    print("\n----- OWNER SUMMARY -----")
    print(f"{'Owner':<20} {'Count':<8} {'Total Size':<20} {'% of Memory':<12} {'PIDs'}")
    print("-" * 90)
   
    # Sort owners by total allocated memory
    sorted_owners = sorted(summary['by_owner'].items(), key=lambda x: x[1]['total_allocated'], reverse=True)
   
    for owner, data in sorted_owners:
        percent = (data['total_allocated'] / summary['total_allocated_memory'] * 100) if summary['total_allocated_memory'] else 0
        pids = ", ".join(sorted(data['pids']))
        print(f"{owner:<20} {data['count']:<8} {data['total_allocated']:<12} ({format_size(data['total_allocated']):<6}) {percent:<12.2f} {pids}")
   
    print("\n----- CRITICAL COMPONENTS -----")
   
    # Print top 5 memory-intensive PIDs
    print("Top memory-intensive PIDs:")
    for i, pid_info in enumerate(summary['top_pids'], 1):
        print(f"{i}. PID {pid_info['pid']} with {pid_info['total_allocated']} bytes allocated ({format_size(pid_info['total_allocated'])})")
   
    # Print top 5 memory-intensive owners
    print("\nTop memory-intensive Owners:")
    for i, owner_info in enumerate(summary['top_owners'], 1):
        pids_count = len(owner_info['pids'])
        pids_text = f"across {pids_count} PID{'' if pids_count == 1 else 's'}"
        print(f"{i}. {owner_info['owner']} with {owner_info['total_allocated']} bytes allocated ({format_size(owner_info['total_allocated'])}) {pids_text}")

def export_to_excel(summary, filename):
    """
    Export summary to an Excel file with three sheets

    Args:
        summary (dict): Summary statistics
        filename (str): Output Excel file name
    """
    try:
        print(f"Exporting summary to {filename}...")
        wb = Workbook()
       
        # Create sheets
        ws_summary = wb.active
        ws_summary.title = "Summary"
       
        ws_pid_summary = wb.create_sheet("PID Summary")
       
        ws_owner_summary = wb.create_sheet("Owner Summary")
       
        # Export overall statistics, top 5 PIDs, top 5 owners, and largest allocation to the Summary sheet
        ws_summary.append(['Metric', 'Value', 'Human Readable'])
        ws_summary.append(['Total failures', summary['total_failures'], ''])
        ws_summary.append(['Total requested memory', summary['total_requested_size'], format_size(summary['total_requested_size'])])
        ws_summary.append(['Average request size', f"{summary['average_requested_size']:.2f}", format_size(summary['average_requested_size'])])
        ws_summary.append(['Total allocated memory', summary['total_allocated_memory'], format_size(summary['total_allocated_memory'])])
       
        if summary['largest_allocation']['size'] > 0:
            ws_summary.append(['', '', ''])
            ws_summary.append(['Largest Allocation', '', ''])
            ws_summary.append(['Size', summary['largest_allocation']['size'], format_size(summary['largest_allocation']['size'])])
            ws_summary.append(['PID', summary['largest_allocation']['pid'], ''])
            ws_summary.append(['Owner', summary['largest_allocation']['owner'], ''])
            ws_summary.append(['Address', summary['largest_allocation']['address'], ''])
       
        # Export top 5 memory-intensive PIDs
        ws_summary.append(['', '', ''])
        ws_summary.append(['Top 5 Memory-Intensive PIDs', '', ''])
        ws_summary.append(['#', 'PID', 'Total Allocated (bytes)', 'Human Readable Size'])
        for i, pid_info in enumerate(summary['top_pids'], 1):
            ws_summary.append([i, pid_info['pid'], pid_info['total_allocated'], format_size(pid_info['total_allocated'])])
       
        # Export top 5 memory-intensive owners
        ws_summary.append(['', '', ''])
        ws_summary.append(['Top 5 Memory-Intensive Owners', '', ''])
        ws_summary.append(['#', 'Owner', 'Total Allocated (bytes)', 'Human Readable Size', '# of PIDs'])
        for i, owner_info in enumerate(summary['top_owners'], 1):
            pids_count = len(owner_info['pids'])
            ws_summary.append([i, owner_info['owner'], owner_info['total_allocated'], format_size(owner_info['total_allocated']), pids_count])
       
        # Export PID summary to the PID Summary sheet
        ws_pid_summary.append(['PID', 'Count', 'Total Allocated (bytes)', 'Human Readable Size', '% of Memory'])
        sorted_pids = sorted(summary['by_pid'].items(), key=lambda x: x[1]['total_allocated'], reverse=True)
       
        for pid, data in sorted_pids:
            percent = (data['total_allocated'] / summary['total_allocated_memory'] * 100) if summary['total_allocated_memory'] else 0
            ws_pid_summary.append([pid, data['count'], data['total_allocated'], format_size(data['total_allocated']), f"{percent:.2f}%"])
       
        # Export Owner summary to the Owner Summary sheet
        ws_owner_summary.append(['Owner', 'Count', 'Total Allocated (bytes)', 'Human Readable Size', '% of Memory', 'PIDs'])
        sorted_owners = sorted(summary['by_owner'].items(), key=lambda x: x[1]['total_allocated'], reverse=True)
       
        for owner, data in sorted_owners:
            percent = (data['total_allocated'] / summary['total_allocated_memory'] * 100) if summary['total_allocated_memory'] else 0
            pids = ", ".join(sorted(data['pids']))
            ws_owner_summary.append([owner, data['count'], data['total_allocated'], format_size(data['total_allocated']), f"{percent:.2f}%", pids])
       
        wb.save(filename)
        print(f"Summary exported to {filename}")
    except Exception as e:
        print(f"Error exporting to Excel: {str(e)}")

def main():
    parser = argparse.ArgumentParser(description='Analyze TizenRT memory allocation failure logs')
    parser.add_argument('logfile', help='Path to the log file')
    parser.add_argument('--export', help='Export summary to excel file')
    args = parser.parse_args()
   
    print(f"\nTizenRT Memory Allocation Failure Log Analyzer")
    print(f"==============================================")
   
    # Get file size
    try:
        file_size = os.path.getsize(args.logfile)
        print(f"Log file size: {format_size(file_size)}")
    except FileNotFoundError:
        print(f"Error: File '{args.logfile}' not found")
        sys.exit(1)
   
    start_time = time.time()
    
    # Always preprocess the log file
    processed_file = preprocess_log_file(args.logfile)
   
    # First pass: count failures and gather file metrics
    total_failures, file_size, failure_positions = count_failures(processed_file)

    # Second pass: process failures and generate summary
    summary = parse_log_file(processed_file, total_failures, file_size, failure_positions)
   
    # Print summary
    print_summary(summary)
   
    # Export if requested
    if args.export:
        export_to_excel(summary, args.export)
   
    # Clean up processed file if it was created
    if processed_file != args.logfile:
        try:
            os.remove(processed_file)
            print(f"Cleaned up temporary processed file.")
        except:
            print(f"Note: Could not remove temporary file: {processed_file}")
   
    total_time = time.time() - start_time
    print(f"\nTotal analysis time: {format_time(total_time)}")
    print("Analysis complete.")

if __name__ == "__main__":
    main()